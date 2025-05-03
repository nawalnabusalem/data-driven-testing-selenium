import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

@pytest.fixture(scope='session')
def driver():
    driver = webdriver.Chrome(service=Service(r"your_driver_path"))
    yield driver
    driver.quit()

@pytest.fixture(autouse=True)
def load_webpage(driver):
    driver.get("https://the-internet.herokuapp.com/login")


def load_login_excel_data():
    data = []

    try:
        sheet = load_workbook("data/login_excel_data.xlsx").active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # skip titles
            data.append({
                "username": row[0],
                "password": row[1],
                "expected_success": row[2]
            })

    except Exception as e:
        print(f'Failed to load login excel data due to the error : {e}')

    return data


@pytest.mark.parametrize("login_data", load_login_excel_data())
def test_login_page(driver, login_data):
    username = login_data["username"]
    password = login_data["password"]
    expected_success = login_data["expected_success"]

    try:

        input_username = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "username"))
        )
        input_password = driver.find_element(by=By.ID, value="password")

        input_username.clear()
        input_password.clear()

        input_username.send_keys(username)
        input_password.send_keys(password)

        driver.find_element(by=By.CLASS_NAME, value="radius").submit()
        login_status_message = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "flash")))

        if expected_success:
            assert "You logged into a secure area!" in login_status_message.text, "Expected successful login for valid test case"

            print(f"\033[32m ✓ PASS: Valid login for username: {username}, password:{password} \033[0m")

            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#content > div > a"))
            ).click()

        else:
            assert 'invalid' in login_status_message.text, "Expected failure for invalid test case"

            print(f"\033[32m ✓ PASS: Invalid credentials is handled successfully for username: {username}, "
                  f"password:{password} \033[0m")


    except Exception as e:
        print(f"\033[33m ⚠ ERROR : Failed Test Case for username: {username}, password:{password} "
              f"\n{e}\033[0m")
